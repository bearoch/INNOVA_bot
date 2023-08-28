import os
from openpyxl import Workbook
from datetime import timedelta, datetime

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from models.methods.select import get_position_history, get_salary_history
from models.models import PositionTab, DurationOfProbation, CityOfResidence, EmployeePosition, \
    Employee, SalaryHistory
from config_data.config import Config, load_config

config: Config = load_config('.env')
engine = create_engine(f'postgresql://{config.db.db_user}:{config.db.db_password}@{config.db.db_host}/{config.db.database}', echo=True)

session = sessionmaker(bind=engine)
s = session()

def create_xls_file(table=''):
    if table == 'all_employs':
        try:
            book = Workbook()
            book.remove(book.active)
            sheet_1 = book.create_sheet("Список всех сотрудников", 0)
            colons = ['id', 'ФИО', 'Дата приема на работу', 'Должность', 'Наличие ИС', 'Продолжительность ИС',
                      "Дата завершения ИС", 'ЗП во время ИС', 'ЗП после ИС', 'Город проживания', 'Телефонный номер',
                      'Дата рождения']
            sheet_1.append(colons)
            new_list: list = []
            for row in s.query(Employee, DurationOfProbation, CityOfResidence).filter(
                        Employee.city_of_residence_id == CityOfResidence.city_of_residence_id,
                            Employee.duration_of_probation_id == DurationOfProbation.duration_of_probation_id):
                if row.Employee.have_duration_of_probation == "Да" and not row.Employee.date_of_dismissal:
                    date_end_duration_of_probation = row.Employee.date_of_employment + timedelta(
                        days=int(row.DurationOfProbation.duration_of_probation))
                    new_list.append([row.Employee.employee_id,
                                    row.Employee.name_employee,
                                    row.Employee.date_of_employment.strftime('%d.%m.%Y'),
                                    max(get_position_history(str(row.Employee.employee_id)).items())[1],
                                    row.Employee.have_duration_of_probation,
                                    row.DurationOfProbation.duration_of_probation,
                                    date_end_duration_of_probation.strftime('%d.%m.%Y'),
                                    row.Employee.salary_on_probation,
                                    max(get_salary_history(str(row.Employee.employee_id)).items())[1],
                                    row.CityOfResidence.name_city,
                                    row.Employee.phone_number,
                                    row.Employee.date_of_birth.strftime('%d.%m.%Y')
                                    ])
            for row1 in s.query(Employee, CityOfResidence).filter(Employee.city_of_residence_id == CityOfResidence.city_of_residence_id):
                if row1.Employee.have_duration_of_probation == "Нет" and not row1.Employee.date_of_dismissal:
                    new_list.append([row1.Employee.employee_id,
                                     row1.Employee.name_employee,
                                     row1.Employee.date_of_employment.strftime('%d.%m.%Y'),
                                     max(get_position_history(str(row1.Employee.employee_id)).items())[1],
                                     row1.Employee.have_duration_of_probation,
                                     '-',
                                     '-',
                                     '-',
                                     max(get_salary_history(str(row1.Employee.employee_id)).items())[1],
                                     row1.CityOfResidence.name_city,
                                     row1.Employee.phone_number,
                                     row1.Employee.date_of_birth.strftime('%d.%m.%Y')
                                     ])
            for x in sorted(new_list, key=lambda x: x[0]):
                sheet_1.append(x)

            print(sorted(new_list, key=lambda x: x[0]))

            book.save("employee.xlsx")

            return os.path.abspath('employee.xlsx')
        except:
            print('Не сформирован список всех сотрудников в Excel')

    elif table == 'dismissal':
        try:
            book = Workbook()
            book.remove(book.active)
            sheet_1 = book.create_sheet("Список уволенных сотрудников", 0)
            colons = ['№', 'ФИО', 'Дата приема на работу', 'Должность', 'Наличие ИС', 'Продолжительность ИС',
                      'ЗП во время ИС', 'ЗП после ИС', 'Город проживания', 'Телефонный номер', 'Дата рождения',
                      'Дата увольнения']
            sheet_1.append(colons)
            new_list: list = []
            for row in s.query(Employee, DurationOfProbation, CityOfResidence).filter(
                        Employee.city_of_residence_id == CityOfResidence.city_of_residence_id,
                            Employee.duration_of_probation_id == DurationOfProbation.duration_of_probation_id):
                if row.Employee.have_duration_of_probation == "Да" and row.Employee.date_of_dismissal:
                    new_list.append([row.Employee.employee_id,
                                    row.Employee.name_employee,
                                    row.Employee.date_of_employment.strftime('%d.%m.%Y'),
                                    max(get_position_history(str(row.Employee.employee_id)).items())[1],
                                    row.Employee.have_duration_of_probation,
                                    row.DurationOfProbation.duration_of_probation,
                                    row.Employee.salary_on_probation,
                                    max(get_salary_history(str(row.Employee.employee_id)).items())[1],
                                    row.CityOfResidence.name_city,
                                    row.Employee.phone_number,
                                    row.Employee.date_of_birth.strftime('%d.%m.%Y'),
                                    row.Employee.date_of_dismissal.strftime('%d.%m.%Y')
                                    ])
            for row1 in s.query(Employee, CityOfResidence).filter(Employee.city_of_residence_id == CityOfResidence.city_of_residence_id):
                if row1.Employee.have_duration_of_probation == "Нет" and row1.Employee.date_of_dismissal:
                    new_list.append([row1.Employee.employee_id,
                                     row1.Employee.name_employee,
                                     row1.Employee.date_of_employment.strftime('%d.%m.%Y'),
                                     max(get_position_history(str(row1.Employee.employee_id)).items())[1],
                                     row1.Employee.have_duration_of_probation,
                                     '-',
                                     '-',
                                     max(get_salary_history(str(row1.Employee.employee_id)).items())[1],
                                     row1.CityOfResidence.name_city,
                                     row1.Employee.phone_number,
                                     row1.Employee.date_of_birth.strftime('%d.%m.%Y'),
                                     row1.Employee.date_of_dismissal.strftime('%d.%m.%Y')
                                     ])
            for x in sorted(new_list, key=lambda x: x[0]):
                sheet_1.append(x)

            print(sorted(new_list, key=lambda x: x[0]))

            book.save("employee.xlsx")

            return os.path.abspath('employee.xlsx')
        except:
            print('Не сформирован список всех сотрудников в Excel')

    elif table == 'have_duration_of_probation':
        try:
            book = Workbook()
            book.remove(book.active)
            sheet_1 = book.create_sheet("Список сотрудников на ИС", 0)
            colons = ['id', 'ФИО', 'Дата приема на работу', 'Должность', 'Продолжительность ИС', 'Дата завершения ИС',
                      'ЗП во время ИС', 'ЗП после ИС', 'Город проживания', 'Телефонный номер', 'Дата рождения']
            sheet_1.append(colons)
            new_list: list = []
            new_list: list = []
            for row in s.query(Employee, DurationOfProbation, CityOfResidence).filter(
                    Employee.city_of_residence_id == CityOfResidence.city_of_residence_id,
                    Employee.duration_of_probation_id == DurationOfProbation.duration_of_probation_id):
                if row.Employee.have_duration_of_probation == "Да" and not row.Employee.date_of_dismissal:
                    finish_probation: datetime = row.Employee.date_of_employment + timedelta(days=row.DurationOfProbation.duration_of_probation)
                    new_list.append([row.Employee.employee_id,
                                     row.Employee.name_employee,
                                     row.Employee.date_of_employment.strftime('%d.%m.%Y'),
                                     max(get_position_history(str(row.Employee.employee_id)).items())[1],
                                     row.DurationOfProbation.duration_of_probation,
                                     finish_probation.strftime('%d.%m.%Y'),
                                     row.Employee.salary_on_probation,
                                     max(get_salary_history(str(row.Employee.employee_id)).items())[1],
                                     row.CityOfResidence.name_city,
                                     row.Employee.phone_number,
                                     row.Employee.date_of_birth.strftime('%d.%m.%Y')
                                     ])

            for x in sorted(new_list, key=lambda x: x[0]):
                sheet_1.append(x)

            book.save("employee.xlsx")

            return os.path.abspath('employee.xlsx')


        except:
            print('Не сформирован список сотрудников находящихся на ИС в Excel')

    elif table == 'position_history':
        try:
            book = Workbook()
            book.remove(book.active)
            sheet_1 = book.create_sheet("История должностей", 0)
            colons = ['№', 'ФИО', 'Должность', 'Дата постановки на должность']
            sheet_1.append(colons)

            new_list: list = []
            for row in s.query(EmployeePosition, Employee, PositionTab).filter(
                    EmployeePosition.employee_id == Employee.employee_id,
                    EmployeePosition.position_id == PositionTab.position_id):
                if not row.Employee.date_of_dismissal:

                    new_list.append([
                                     row.Employee.name_employee,
                                     row.PositionTab.name_position,
                                     row.EmployeePosition.date_of_change.strftime('%d.%m.%Y')
                                     ])

            for index, x in enumerate(sorted(new_list, key=lambda x: x[0]),1):
                sheet_1.append([index] + x)

            book.save("employee.xlsx")

            return os.path.abspath('employee.xlsx')


        except:
            print('Не сформирован список истории должностей сотрудников в Excel')

    elif table == 'salary_history':
        try:
            book = Workbook()
            book.remove(book.active)
            sheet_1 = book.create_sheet("История зарплат", 0)
            colons = ['№', 'ФИО', 'Объем ЗП', 'Дата']
            sheet_1.append(colons)

            new_list: list = []
            for row in s.query(SalaryHistory, Employee).filter(SalaryHistory.employee_id == Employee.employee_id):
                if not row.Employee.date_of_dismissal:
                    new_list.append([
                                     row.Employee.name_employee,
                                     row.SalaryHistory.salary,
                                     row.SalaryHistory.date_of_change.strftime('%d.%m.%Y')
                                     ])

            for index, x in enumerate(sorted(new_list, key=lambda x: x[0]), 1):
                sheet_1.append([index] + x)

            book.save("employee.xlsx")

            return os.path.abspath('employee.xlsx')


        except:
            print('Не сформирован список истории должностей сотрудников в Excel')

